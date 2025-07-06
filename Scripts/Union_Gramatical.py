import openpyxl
import pandas as pd
import re
from tkinter import filedialog, Tk
import spacy
from collections import defaultdict
import time

nlp = spacy.load("es_core_news_sm")  # Carga el modelo en español

# Crear una ventana emergente para que el usuario seleccione el archivo de texto
root = Tk()
root.withdraw()  # Ocultar la ventana principal

# Pedir al usuario que seleccione el archivo de texto
selected_file = filedialog.askopenfilename(filetypes=[("Archivos de texto", "*.txt")])
if not selected_file:
    print("No se seleccionó ningún archivo.")
else:
    # Cargar el archivo de texto con la transcripción de la batalla de rap
    with open(selected_file, 'r', encoding='utf-8') as file:
        texto = file.read()

    # Dividir el texto en palabras utilizando expresiones regulares
    palabras = re.findall(r'\w+', texto.lower())  # Convertir a minúsculas

    # Contar la frecuencia de cada palabra
    frecuencia_palabras = pd.Series(palabras).value_counts()

    # Crear un DataFrame para almacenar los resultados
    df = pd.DataFrame({'Palabra': frecuencia_palabras.index, 'Frecuencia': frecuencia_palabras.values})

    # Lista de patrones de terminación
    patrones = ["ar", "ro", "mo", "er", "ir", "or", "al", "ista", "ción", "ista", "ón", "ble", "aje", "ista", "ible"]

    # Etiquetar gramaticalmente las palabras y escribir las familias en una nueva columna "Familia"
    df['Etiqueta_Gramatical'] = df['Palabra'].apply(lambda palabra: nlp(palabra)[0].pos_ if nlp(palabra) and len(nlp(palabra)) > 0 else "Desconocida")

    # Reordenar las columnas en el DataFrame
    df = df[['Etiqueta_Gramatical', 'Palabra', 'Frecuencia']]

    # Agregar columnas para los patrones
    for patron in patrones:
        df[patron] = df['Palabra'].apply(lambda palabra: palabra.endswith(patron))

    # Pedir al usuario que ingrese el nombre del archivo de salida
    output_file_name = input("Ingresa el nombre: ")

    # Construir la ruta completa del archivo de salida
    output_file = f"C:\\Users\\PEPE\\Documents\\Personal\\MCP\\RedBullBatallas-20-años\\excel\\{output_file_name}.xlsx"

    # Crear un archivo Excel con los resultados
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Agregar cabeceras en la fila 1
    cabeceras = list(df.columns)
    for col, cabecera in enumerate(cabeceras, 1):
        sheet.cell(row=1, column=col, value=cabecera)

    # Copiar los datos de las columnas "Palabra" y "Frecuencia" en las columnas B y A
    for row, (familia, palabra, frecuencia) in enumerate(zip(df['Etiqueta_Gramatical'], df['Palabra'], df['Frecuencia']), start=3):
        sheet.cell(row=row, column=1, value=familia)
        sheet.cell(row=row, column=2, value=palabra)
        sheet.cell(row=row, column=3, value=frecuencia)

    # Copiar los valores booleanos de los patrones a las columnas restantes
    for patron in patrones:
        for row, valor in enumerate(df[patron], start=3):
            sheet.cell(row=row, column=cabeceras.index(patron) + 1, value=valor)

    # Contar el número de celdas en la columna B
    num_celdas_B = len(df)
    sheet.cell(row=2, column=2, value=num_celdas_B)

    # Sumar los valores en la columna B (Frecuencia)
    suma_celdas_B = df['Frecuencia'].sum()
    sheet.cell(row=2, column=3, value=suma_celdas_B)

    # Sumar los valores en las columnas D a P
    for col in range(4, 17):  # Columnas D a P
        suma_columna = df.iloc[:, col - 1].sum()
        sheet.cell(row=2, column=col, value=suma_columna)

    
    # Guardar el archivo Excel

    workbook.save(output_file)
    time.sleep(3)

    #Nueva logica para integrar 

    # Cargar el archivo Excel
    archivo_excel = output_file

    # Crear una nueva hoja llamada 'Conteo_Familias'
    sheet_conteo = workbook.create_sheet('Conteo_Familias')

    # Crear un diccionario para almacenar las palabras por familia
    palabras_por_familia = defaultdict(list)

    # Obtener los valores de la columna 'Familia' y 'Palabra'
    valores_familia = [cell.value for cell in sheet['A'][2:]]  # Excluyendo la primera celda que es la etiqueta
    valores_palabra = [cell.value for cell in sheet['B'][2:]]  # Excluyendo la primera celda que es la etiqueta

    # Agrupar las palabras por familia
    for familia, palabra in zip(valores_familia, valores_palabra):
        palabras_por_familia[familia].append(palabra)

    # Crear nuevas columnas para cada familia
    nueva_columna = sheet_conteo.max_column + 1

    for idx, (familia, palabras) in enumerate(palabras_por_familia.items(), start=nueva_columna):
        sheet_conteo.cell(row=1, column=idx, value=f"{familia}")
        for row, palabra in enumerate(palabras, start=3):
            sheet_conteo.cell(row=row, column=idx, value=palabra)

    # Contar las celdas no vacías en cada columna y colocar los resultados en la fila 2
    for col in range(1, sheet_conteo.max_column + 1):
        count_value = len([cell.value for cell in sheet_conteo[openpyxl.utils.get_column_letter(col)][2:] if cell.value is not None])
        sheet_conteo.cell(row=2, column=col, value=count_value)


    # Obtener las 10 palabras más repetidas con más de 3 letras
    palabras_mas_repetidas = df[(df['Palabra'].str.len() > 3)].nlargest(25, 'Frecuencia')

    # Obtener las 5 terminaciones más usadas
    terminaciones_mas_usadas = df.loc[:, 'ar':'ible'].sum().nlargest(10)

    # Crear una nueva hoja llamada 'Palabras_Mas_Repetidas_y_Terminaciones'
    sheet_palabras_terminaciones = workbook.create_sheet('Frecuencia_y_Terminaciones')

    # Agregar las palabras más repetidas a la hoja
    sheet_palabras_terminaciones.cell(row=1, column=1, value='Palabras Más Repetidas')
    for row, (palabra, frecuencia) in enumerate(zip(palabras_mas_repetidas['Palabra'], palabras_mas_repetidas['Frecuencia']), start=2):
        sheet_palabras_terminaciones.cell(row=row, column=1, value=f"{palabra} - {frecuencia}")

    # Agregar las terminaciones más usadas a la hoja
    sheet_palabras_terminaciones.cell(row=1, column=3, value='Terminaciones Más Usadas')
    for row, (terminacion, frecuencia) in enumerate(terminaciones_mas_usadas.items(), start=2):
        sheet_palabras_terminaciones.cell(row=row, column=3, value=f"{terminacion} - {frecuencia}")



    
    # Crear una nueva hoja llamada 'Conteo_Familias'
    sheet_conteo = workbook.create_sheet('Vertical')

    # Leer el contenido del archivo de texto
    with open(selected_file, 'r', encoding='utf-8') as file:
        # Convertir el texto a una lista de palabras
        palabras = file.read().split()

        # Iterar sobre las palabras
        for idx, palabra in enumerate(palabras):
            # Escribir la palabra en la celda correspondiente
            sheet_conteo.cell(row=idx + 2, column=4, value=palabra)

    # Guardar el archivo Excel
    workbook.save(archivo_excel)


    print(f"Terminó el conteo y análisis || El archivo Excel se guardó")

    try:
        root.destroy()  # Intenta cerrar la ventana emergente
    except Exception as e:
        print(f"Error al cerrar la ventana: {e}")

